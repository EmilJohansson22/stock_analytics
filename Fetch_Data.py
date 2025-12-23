import yfinance as yf
import pandas as pd
from datetime import datetime
from openpyxl import load_workbook
from openpyxl.styles import NamedStyle  # Optional, for custom styles if needed

def fetch_and_save(ticker):
    """
    Fetches specified data for a ticker from Yahoo Finance and appends it to data_population.xlsx.
    Creates the file if it doesn't exist.
    """
    try:
        # Fetch ticker data
        t = yf.Ticker(ticker)
        info = t.info
        
        # Quarterly statements (for TTM and latest BS)
        q_is = t.quarterly_income_stmt
        q_cf = t.quarterly_cashflow
        q_bs = t.quarterly_balance_sheet
        
        # Annual statements (for oldest revenue)
        a_is = t.financials
        
        # Basic info
        currency = info.get('currency', 'N/A')
        nationality = info.get('country', 'N/A')
        # Diluted Average Shares from latest quarterly income statement, fallback to Basic, then to current shares
        shares_outstanding = None
        shares_type = None
        if not q_is.empty:
            for key in ['Diluted Average Shares', 'Weighted Average Diluted Shares Outstanding', 'Diluted Shares Outstanding', 'Basic Average Shares', 'Weighted Average Basic Shares Outstanding', 'Basic Shares Outstanding']:
                if key in q_is.index:
                    shares_outstanding = q_is.loc[key, q_is.columns[0]]
                    shares_type = key
                    break
        if shares_outstanding is None:
            shares_outstanding = info.get('sharesOutstanding')
            shares_type = "Current Shares Outstanding"
        
        # TTM calculations (sum last 4 quarters if available)
        def sum_ttm(series, key_variants):
            if series.empty:
                return None
            # Find the column with the data
            for variant in key_variants:
                if variant in series.index:
                    vals = series.loc[variant, series.columns[:4]].dropna()
                    return vals.sum() if len(vals) > 0 else None
            return None
        
        ebit_ttm = sum_ttm(q_is, ['Operating Income', 'EBIT', 'Operating income'])  # Assuming EBITA means EBIT
        ebitda_ttm = sum_ttm(q_is, ['EBITDA'])
        net_income_ttm = sum_ttm(q_is, ['Net Income'])
        revenue_ttm = sum_ttm(q_is, ['Total Revenue', 'Revenue'])
        cost_of_revenue_ttm = sum_ttm(q_is, ['Cost Of Revenue', 'Cost of Revenue'])
        operating_expenses_ttm = sum_ttm(q_is, ['Operating Expense', 'Operating Expenses'])
        tax_provision_ttm = sum_ttm(q_is, ['Tax Provision', 'Income Tax Expense'])
        depreciation_amortization_ttm = sum_ttm(q_cf, ['Depreciation And Amortization', 'Depreciation'])
        capital_expenditures_ttm = sum_ttm(q_cf, ['Capital Expenditures', 'Capital Expenditure', 'Capex', 'Property, Plant and Equipment', 'Additions to Property, Plant and Equipment'])
        # Ensure CapEX is positive (outflow)
        if capital_expenditures_ttm is not None:
            capital_expenditures_ttm = abs(capital_expenditures_ttm)
        
        # Change in Working Capital (from cash flow, TTM sum)
        change_wc_ttm = sum_ttm(q_cf, ['Change In Working Capital'])
        # Ensure Change in WC is positive (outflow magnitude)
        if change_wc_ttm is not None:
            change_wc_ttm = abs(change_wc_ttm)
        
        # Revenue TTM Date: Date of latest quarterly report
        revenue_ttm_date = q_is.columns[0] if not q_is.empty else None
        
        # Oldest Annual Revenue and Date
        oldest_annual_revenue = None
        oldest_revenue_date = None
        if not a_is.empty:
            oldest_col = a_is.columns[-1]  # Last column is oldest
            if 'Total Revenue' in a_is.index:
                oldest_annual_revenue = a_is.loc['Total Revenue', oldest_col]
            oldest_revenue_date = oldest_col
        
        # Balance Sheet (latest quarterly)
        latest_bs = q_bs.iloc[:, 0] if not q_bs.empty else pd.Series(dtype=float)
        shareholders_equity = latest_bs.get('Total Stockholder Equity') or latest_bs.get('Shareholders Equity')
        total_assets = latest_bs.get('Total Assets')
        total_debt = latest_bs.get('Total Debt') or (latest_bs.get('Long Term Debt') + latest_bs.get('Short Term Debt') if pd.notna(latest_bs.get('Long Term Debt')) and pd.notna(latest_bs.get('Short Term Debt')) else None)
        # Cash and Short Term Investments
        cash_equivalents = latest_bs.get('Cash And Cash Equivalents', 0) or latest_bs.get('Cash', 0)
        short_term_investments = latest_bs.get('Short Term Investments', 0)
        cash_and_equivalents = cash_equivalents + short_term_investments
        
        # Prepare data dict
        data = {
            'Ticker': ticker,
            'Currency': currency,
            'Nationality': nationality,
            'Shares_Outstanding': shares_outstanding,
            'EBITA': ebit_ttm,  # Assuming EBITA means EBIT
            'EBITDA': ebitda_ttm,
            'Net_Income': net_income_ttm,
            'Revenue_TTM': revenue_ttm,
            'Revenue_TTM_Date': revenue_ttm_date,
            'Oldest_Annual_Revenue': oldest_annual_revenue,
            'Oldest_Revenue_Date': oldest_revenue_date,
            'Cost_of_Revenue': cost_of_revenue_ttm,
            'Operating_Expenses': operating_expenses_ttm,
            'Tax_Provision': tax_provision_ttm,
            'Depreciation_Amortization': depreciation_amortization_ttm,
            'Capital_Expenditures': capital_expenditures_ttm,
            'Shareholders_Equity': shareholders_equity,
            'Total_Assets': total_assets,
            'Total_Debt': total_debt,
            'Cash_and_Equivalents': cash_and_equivalents,
            'Change_in_Working_Capital': change_wc_ttm,
            'Fetch_Date': datetime.now().strftime('%Y-%m-%d %H:%M:%S')
        }
        
        # Load or create Excel file
        file_path = 'data_population.xlsx'
        try:
            existing_df = pd.read_excel(file_path)
        except FileNotFoundError:
            existing_df = pd.DataFrame()
        
        # Append new row
        new_df = pd.DataFrame([data])
        updated_df = pd.concat([existing_df, new_df], ignore_index=True)
        
        # Save back to Excel
        updated_df.to_excel(file_path, index=False)
        print(f"Data for {ticker} appended to {file_path}")
        
    except Exception as e:
        print(f"Error fetching data for {ticker}: {e}")

def process_excel_file(file_path='data_population.xlsx'):
    """
    Reads data_population.xlsx (Row 1: headers, Row 2+: data), fetches data for each ticker in the 'Yahoo Ticker' column,
    updates the specified columns, and writes back to the file.
    Adds a 'Notes' column for Change_in_Working_Capital method and shares type.
    Formats numeric cells to display as plain numbers (no commas), and Tax_Rate as percentage.
    Auto-adjusts column widths based on content.
    """
    try:
        # Read the Excel file, treating Row 1 as headers
        df = pd.read_excel(file_path, header=0)
        # No drop needed since Row 1 is headers, Row 2+ is data
        
        ticker_column = 'Yahoo Ticker'  # Use 'Yahoo Ticker' as the column for tickers
        if ticker_column not in df.columns:
            print(f"Error: '{ticker_column}' column not found in the Excel file.")
            return
        
        # Columns to update (using the exact names from the file)
        update_columns = [
            'Currency', 'Nationality', 'Shares_Outstanding', 'EBIT', 'EBITA', 'EBITDA', 'Net_Income',
            'Revenue TTM', 'Revenue TTM Date', 'Oldest Revenue Yahoo Finance', 'Oldest Revenue Year',
            'COGS', 'Operating_Expenses', 'Tax_Rate', 'Depreciation_Amortization', 'Capital_Expenditures',
            'Shareholders Equity', 'Total_Assets', 'Debt', 'Cash', 'Change_in_Working_Capital', 'Notes'
        ]
        
        # Process each row
        for idx, row in df.iterrows():
            ticker = row.get(ticker_column)
            if pd.isna(ticker) or not ticker:
                print(f"Skipping row {idx + 2}: No ticker found.")  # +2 since Row 2 is idx 0
                continue
            
            print(f"Fetching data for {ticker}...")
            try:
                # Fetch data (similar to previous fetch_and_save logic)
                t = yf.Ticker(ticker)
                info = t.info
                q_is = t.quarterly_income_stmt
                q_cf = t.quarterly_cashflow
                q_bs = t.quarterly_balance_sheet
                a_is = t.financials
                
                # Basic info
                currency = info.get('currency', 'N/A')
                nationality = info.get('country', 'N/A')
                # Diluted Average Shares from latest quarterly income statement, fallback to Basic, then to current shares
                shares_outstanding = None
                shares_type = None
                if not q_is.empty:
                    for key in ['Diluted Average Shares', 'Weighted Average Diluted Shares Outstanding', 'Diluted Shares Outstanding', 'Basic Average Shares', 'Weighted Average Basic Shares Outstanding', 'Basic Shares Outstanding']:
                        if key in q_is.index:
                            shares_outstanding = q_is.loc[key, q_is.columns[0]]
                            shares_type = key
                            break
                if shares_outstanding is None:
                    shares_outstanding = info.get('sharesOutstanding')
                    shares_type = "Current Shares Outstanding"
                
                # TTM calculations
                def sum_ttm(series, key_variants):
                    if series.empty:
                        return None
                    for variant in key_variants:
                        if variant in series.index:
                            vals = series.loc[variant, series.columns[:4]].dropna()
                            return vals.sum() if len(vals) > 0 else None
                    return None
                
                ebit_ttm = sum_ttm(q_is, ['Operating Income', 'EBIT', 'Operating income'])
                ebitda_ttm = sum_ttm(q_is, ['EBITDA'])
                net_income_ttm = sum_ttm(q_is, ['Net Income'])
                revenue_ttm = sum_ttm(q_is, ['Total Revenue', 'Revenue'])
                cost_of_revenue_ttm = sum_ttm(q_is, ['Cost Of Revenue', 'Cost of Revenue'])
                operating_expenses_ttm = sum_ttm(q_is, ['Operating Expense', 'Operating Expenses'])
                tax_provision_ttm = sum_ttm(q_is, ['Tax Provision', 'Income Tax Expense'])
                depreciation_amortization_ttm = sum_ttm(q_cf, ['Depreciation And Amortization', 'Depreciation'])
                capital_expenditures_ttm = sum_ttm(q_cf, ['Capital Expenditures', 'Capital Expenditure', 'Capex', 'Property, Plant and Equipment', 'Additions to Property, Plant and Equipment'])
                # Ensure CapEX is positive (outflow)
                if capital_expenditures_ttm is not None:
                    capital_expenditures_ttm = abs(capital_expenditures_ttm)
                
                # Change in Working Capital: Try cash flow first, else estimate
                change_wc_ttm = sum_ttm(q_cf, ['Change In Working Capital'])
                if change_wc_ttm is not None:
                    note = "From Cash Flow"
                    # Ensure Change in WC is positive (outflow magnitude)
                    change_wc_ttm = abs(change_wc_ttm)
                else:
                    # Estimate as 5% of Revenue TTM
                    change_wc_ttm = 0.05 * revenue_ttm if revenue_ttm is not None else None
                    note = "Estimated as 5% of Revenue TTM" if revenue_ttm is not None else "Estimated as None (no Revenue TTM data)"
                    # Estimated is already positive
                
                # Append shares type to note
                note += f"; Shares: {shares_type}"
                
                # Revenue TTM Date: Date of latest quarterly report (date only, no time)
                revenue_ttm_date = q_is.columns[0] if not q_is.empty else None
                if isinstance(revenue_ttm_date, pd.Timestamp):
                    revenue_ttm_date = revenue_ttm_date.date()
                
                # Oldest Annual Revenue and Year (find oldest with data, date only)
                oldest_annual_revenue = None
                oldest_revenue_year = None
                if not a_is.empty:
                    for col in reversed(a_is.columns):  # Start from oldest
                        for key in ['Total Revenue', 'Revenue', 'Revenues']:
                            if key in a_is.index:
                                val = a_is.loc[key, col]
                                if pd.notna(val):
                                    oldest_annual_revenue = val
                                    oldest_revenue_year = col.date() if isinstance(col, pd.Timestamp) else col
                                    break
                        if oldest_annual_revenue is not None:
                            break
                    if oldest_annual_revenue is None:
                        print(f"Debug: No revenue data found in annuals for {ticker}.")
                
                # Balance Sheet
                latest_bs = q_bs.iloc[:, 0] if not q_bs.empty else pd.Series(dtype=float)
                # Try multiple keys for shareholders equity
                shareholders_equity = None
                for key in ['Total Stockholder Equity', 'Stockholders Equity', 'Shareholders Equity', 'Equity', 'Total Equity']:
                    shareholders_equity = latest_bs.get(key)
                    if shareholders_equity is not None:
                        break
                if shareholders_equity is None:
                    print(f"Debug: No shareholders equity key found in BS for {ticker}. Available keys: {list(latest_bs.index)}")
                total_assets = latest_bs.get('Total Assets')
                total_debt = latest_bs.get('Total Debt') or (
                    latest_bs.get('Long Term Debt') + latest_bs.get('Short Term Debt')
                    if pd.notna(latest_bs.get('Long Term Debt')) and pd.notna(latest_bs.get('Short Term Debt'))
                    else None
                )
                # Cash and Short Term Investments
                cash_equivalents = latest_bs.get('Cash And Cash Equivalents', 0) or latest_bs.get('Cash', 0)
                short_term_investments = latest_bs.get('Short Term Investments', 0)
                cash_and_equivalents = cash_equivalents + short_term_investments
                
                # Tax Rate: Calculate as Tax Provision / EBIT if EBIT > 0
                tax_rate = (tax_provision_ttm / ebit_ttm) if (tax_provision_ttm and ebit_ttm and ebit_ttm != 0) else None
                
                # Update the row with fetched data
                df.at[idx, 'Currency'] = currency
                df.at[idx, 'Nationality'] = nationality
                df.at[idx, 'Shares_Outstanding'] = shares_outstanding
                df.at[idx, 'EBIT'] = ebit_ttm
                df.at[idx, 'EBITA'] = ebitda_ttm  # Set EBITA to EBITDA since they are usually the same
                df.at[idx, 'EBITDA'] = ebitda_ttm
                df.at[idx, 'Net_Income'] = net_income_ttm
                df.at[idx, 'Revenue TTM'] = revenue_ttm
                df.at[idx, 'Revenue TTM Date'] = revenue_ttm_date
                df.at[idx, 'Oldest Revenue Yahoo Finance'] = oldest_annual_revenue
                df.at[idx, 'Oldest Revenue Year'] = oldest_revenue_year
                df.at[idx, 'COGS'] = cost_of_revenue_ttm
                df.at[idx, 'Operating_Expenses'] = operating_expenses_ttm
                df.at[idx, 'Tax_Rate'] = tax_rate
                df.at[idx, 'Depreciation_Amortization'] = depreciation_amortization_ttm
                df.at[idx, 'Capital_Expenditures'] = capital_expenditures_ttm
                df.at[idx, 'Shareholders Equity'] = shareholders_equity
                df.at[idx, 'Total_Assets'] = total_assets
                df.at[idx, 'Debt'] = total_debt
                df.at[idx, 'Cash'] = cash_and_equivalents
                df.at[idx, 'Change_in_Working_Capital'] = change_wc_ttm
                df.at[idx, 'Notes'] = note
                
                print(f"Updated data for {ticker}.")
                
            except Exception as e:
                print(f"Error fetching data for {ticker} in row {idx + 2}: {e}")  # +2 since Row 2 is idx 0
                continue
        
        # Write back to Excel
        df.to_excel(file_path, index=False)
        
        # Format numeric cells: plain numbers for most, percentage for Tax_Rate
        wb = load_workbook(file_path)
        ws = wb.active
        # Find Tax_Rate column
        tax_rate_col = None
        for col_num in range(1, ws.max_column + 1):
            if ws.cell(1, col_num).value == 'Tax_Rate':
                tax_rate_col = col_num
                break
        # Format cells
        for row in ws.iter_rows(min_row=2):  # Skip header row
            for cell in row:
                if isinstance(cell.value, (int, float)):
                    if cell.column == tax_rate_col:
                        cell.number_format = '0%'  # Percentage format
                    else:
                        cell.number_format = '0'  # Plain number format
        
        # Auto-adjust column widths
        for col in ws.columns:
            max_length = 0
            column = col[0].column_letter  # Get the column name
            for cell in col:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = (max_length + 2)
            ws.column_dimensions[column].width = adjusted_width
        
        wb.save(file_path)
        
        print(f"Updated {file_path} with fetched data.")
        
    except Exception as e:
        print(f"Error processing the Excel file: {e}")

# Example usage (uncomment to test)
# fetch_and_save('AAPL')
# fetch_and_save('GOOG')

# Run the function
if __name__ == "__main__":
    process_excel_file()